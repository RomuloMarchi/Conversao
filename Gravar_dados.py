from gettext import install

#from extracao_dados import ExtrairDados
import pandas as pd 
from openpyxl import Workbook, load_workbook

from tkinter import messagebox

#from Localizar_arquivos import LocalizarArquivos
#from processamento import processamento
#from app import MyGUI

class GravarDados:

    def DefinirExcel(self, excel_path):
        '''
        Define em qual arquivo será gravado os dados do PDF '''
        #

        #Atribuindo uma váriavel a Database 
        #Instancia a classe MyGui do Modulo app.py
        #app =  MyGUI()

        #Recebe o caminho do excel escolhido pelo usuário
        self.caminho_excel = excel_path
        #Lê o arquivo em excel com a biblioteca Pandas
        self.arquivo_excel = pd.read_excel(self.caminho_excel)
        #arquivo_excel = pd.read_excel('C:\\Users\marchrom\Documents\project_build1-main\excel_sample\GSD_Alignment_TT.xlsx')
        #arquivo_excel = caminho_excel
        print(self.arquivo_excel)

        # #Atribuindo uma váriavel a coluna de Códigos
        self.codigos_excel = self.arquivo_excel.iloc[:,0]
        
        # #Atribuindo uma váriavel a coluna de Descrições
        self.descricao_excel = self.arquivo_excel.iloc[:,1]

        # #Atribuindo uma váriavel a coluna de SMVs
        self.smv_excel = self.arquivo_excel.iloc[:,2]

        #print(codigos_excel,descricao_excel,smv_excel)
        #return self.arquivo_excel, self.codigos_excel, self.descricao_excel, self.smv_excel, self.caminho_excel
     
    #arquivo_excel = (DefinirExcel())
    
    #arquivo_excel, codigos_excel, descricao_excel, smv_excel, caminho_excel = DefinirExcel()
   
    #print( arquivo_excel)
    #print(codigos_excel,descricao_excel,smv_excel)
    #Definindo uma variavel para a planilha do arquivo excel        
    #planilha = load_workbook(caminho_excel)
    planilha = load_workbook('C:\\Users\marchrom\Documents\Old\PDF_Extract-main\excel_sample\GSD_Alignment_TT.xlsx')

    aba_ativa = planilha.active



    def GravarDados(self,lista_de_codigos, lista_descricoes, lista_de_smv, codigos_excel, descricao_excel, smv_excel, aba_ativa):
        a = 0
        b = 0
        c = 0


        for codigos in len(codigos_excel):
            aba_ativa[f"A{codigos}"] = lista_de_codigos[a]
            codigos += 1
        
        for descricoes in len(descricao_excel):
            aba_ativa[f"B{descricoes}"] = lista_descricoes[b]
            descricoes += 1

    
        for smv in len(smv_excel):
            aba_ativa[f"C{smv}"] = lista_de_smv[c]
            smv += 1
    
   # messagebox.showinfo(tittle=None, message= 'Conversão realizada!')

    
    #Grava e substitui o arquivo selecionado como DB.
  
    #planilha.save(caminho_excel)
    
    
    # print(srcCod)
    # print(srcDesc)
    # print(srcSmv)
    # print(pCod)

    
